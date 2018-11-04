from openpyxl import Workbook, load_workbook
import requests, os, xlrd, re

def xls_to_xlsx(fname):
    xlsBook = xlrd.open_workbook(filename=fname)
    wb = Workbook()

    for i in range(0, xlsBook.nsheets):
        xlsSheet = xlsBook.sheet_by_index(i)
        sheet = wb.active if i == 0 else wb.create_sheet()
        sheet.title = xlsSheet.name

        for row in range(0, xlsSheet.nrows):
            for col in range(0, xlsSheet.ncols):
                sheet.cell(row=row + 1, column=col + 1).value = xlsSheet.cell_value(row, col)
    return wb


def get_type_book(s):
    s = s.lower()
    if re.search("пособии", s) or re.search("пособие", s) or re.search("пособия", s):
        res = 'пособие'
    elif re.search("учебник", s) or re.search("учебнике", s) or re.search("учебника", s):
        res = 'учебник'
    elif re.search("практикум", s) or re.search("практикуме", s) or re.search("практикума",s):
        res = 'практикум'
    elif re.search("монография", s) or re.search("монографию", s) or re.search("монографии",s):
        res = 'монография'
    elif re.search("книга", s) or re.search("книгу", s) or re.search("книги",s) or re.search("книге",s):
        res = 'книга'
    else:
        res = None
    return res

def get_lit_urait_json(search_str):
    #download file
    fname = '/tmp/books_urait.xls'
    f=open(fname,'wb')
    ufr = requests.get('https://biblio-online.ru/search?query={0}&excel=1&all=1'.format(search_str))
    f.write(ufr.content)
    f.close()

    wb = xls_to_xlsx(fname)
    ws = wb.active

    result = []

    for i in range(2,ws.max_row):
        name = ws.cell(row=i, column=2).value
        author = ws.cell(row=i, column=3).value
        place_ed = ws.cell(row=i, column=4).value
        url_book = ws.cell(row=i, column=6).value
        year = int(ws.cell(row=i, column=7).value)
        pages = int(ws.cell(row=i, column=8).value)
        isbn = ws.cell(row=i, column=9).value
        result.append({'name': name, 'author': author, 'place': place_ed, 'isbn': isbn, 'URL': url_book, 'year': year, 'pages': pages, 'type': get_type_book(name)})

    wb.close()
    os.remove(fname)

    return result

def get_lit_lanbook_json(search_str):

    fname = '/tmp/books_lanbook.xlsx'
    f = open(fname, 'wb')
    ufr = requests.get('https://lanbook.com/export/?q={0}&type=excel'.format(search_str))  #загружаем файл xlsx

    f.write(ufr.content)
    f.close()

    wb = load_workbook(fname, data_only=True)  # загружает файл xlsx
    ws = wb.active #активируем первый лист

    result = []

    for i in range(10, ws.max_row):
        if str(ws.cell(row=i, column=1).value).isdigit():
            author = ws.cell(row=i, column=5).value
            name = ws.cell(row=i, column=6).value
            isbn = ws.cell(row=i, column=8).value
            place_ed = ws.cell(row=i, column=11).value
            year = int(ws.cell(row=i, column=12).value)
            pages = int(re.findall(r"(\d+)",str(ws.cell(row=i, column=13).value))[0])
            annotation = str(ws.cell(row=i, column=18).value).lower()

            result.append({'name': name, 'author': author, 'place': place_ed, 'isbn': isbn, 'year': year, 'URL': 'https://lanbook.com',
                           'pages': pages, 'type': get_type_book(annotation)})

    wb.close()
    os.remove(fname)
    return result