from openpyxl import load_workbook
from enum import Enum
import re
from django.utils.timezone import datetime
from .core_funcs import isNone,getSemestrs_2d,remove_dublicates,remove_quotes
from .models import Plans, Directions, Profiles, Discipline, Ministerstvo, Univercity

class TypeEduPlan(Enum):
    FULLTIME = 0    #очная форма обучения
    EXTRAMURAL = 1  #заочная форма обучения
    PART_TIME = 2   #заочно-ускоренная форма обучения


class PlanImport(object):#Класс для импорта учебного плана для очной/заочной/заочно-ускоренной форм обучения в формате xlsx

    def __init__(self, fname, type=TypeEduPlan.FULLTIME, num_prof=1, num_prikaz=1, date_prikaz=datetime.today, departament="", isUpdate=False):
        self.fname = fname      #имя файла для учебного плана
        self.data_title = []    #Данные из заголовка учебного плана
        self.type_edu = type    #Определяет форму обучения
        self.num_profiles = num_prof #Количество профилей в учебном плане
        self.ws = load_workbook(self.fname, data_only=True).active  #загружает файл xlsx
        self.num_prikaz = num_prikaz
        self.date_prikaz = date_prikaz
        self.departament = departament #название кафедры за которой закрепляется направление
        self.isUpdate = isUpdate #Включение/выключение режима обновления учебных планов
        self.header_rows_max = 0 #номер строки в которой заканчивается поиск данных заголовка
        self.rows = [0,0] # Диапазон строк в которых содержатся информация о дисциплинах и часах, компетенциях и т.д.

        for j in range(1, self.ws.max_column):
         for i in range(1, self.ws.max_row):
            s = str(self.ws.cell(row=i, column=j).value).lower()
            if re.search("план\s*учебного\s*процесса", s):
                print("План учебного процесса col=",j," row=",i)
                self.rows[0] = i
            elif re.search("количество\s*экзаменов/зачетов/курсовых", s) or re.search("кол-во\s*экзаменов/зачетов/курсовых", s):
                print("Конец плана col=",j," row=",i)
                self.rows[1] = i
            elif re.search("график\s*учебного\s*процесса", s) or re.search("заочная\s*форма\s*обучения", s): #
                print("График учебного процесса col=",j," row=",i)
                self.header_rows_max = i
            elif re.search("наименование\s*дисциплины", s):
                self.col_name_discipline = j
                print("Номер столбца с дисциплиной col=", j)

        for i in range(1, self.header_rows_max):
            for j in range(1, self.ws.max_column):
                tmp = self.ws.cell(row=i, column=j).value
                if tmp:
                    self.data_title.append(tmp)

        if self.type_edu==TypeEduPlan.FULLTIME:
            print("choise FULLTIME")
        elif self.type_edu==TypeEduPlan.EXTRAMURAL or self.type_edu==TypeEduPlan.PART_TIME:
            self.rows = [self.header_rows_max+1, self.ws.max_row]
            print("choise {0}".format("EXTRAMURAL" if self.type_edu==TypeEduPlan.EXTRAMURAL else "PART-TIME"))
        else:
            print("choise unknown plan")
            return -1

        self.ministerstvo = self.get_ministerstvo()
        self.univercity = self.get_univercity()

    def get_title_data(self):
        return self.data_title


    def get_ministerstvo(self):#плдучить название министерства из учебного плана
        res = ""
        for i in range(len(self.data_title)):  # get direction
            s = str(self.data_title[i])
            if re.search(r'министерство', s.lower()):
                res = s.lstrip()
        return res

    def get_univercity(self): #получить название университета из учебного плана
        res = ""
        for i in range(len(self.data_title)):  # get direction
            s = str(self.data_title[i])
            if re.search(r'учреждение', s.lower()):
                res = s.lstrip()
        return res

    def get_direction(self):
        res = []
        for i in range(len(self.data_title)):#get direction
            s = str(self.data_title[i])
            if re.search(r'НАПРАВЛЕНИЕ:',s):
                tmp = s.split(":")[1]
                code = re.findall(r"\d[0-9].\d[0-9].\d[0-9]",tmp)[0]
                name = re.split(r"\d[0-9].\d[0-9].\d[0-9]", tmp)[1]
                res.append(code)
                res.append(name.lstrip().capitalize())
        return res

    def get_year(self):#извлечение года набора
        res = 2000
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'набор', s.lower()) or re.search(r'год\s*начала\s*подготовки\s*по\s*учебному\s*плану', s.lower()):
                res = int(re.findall(r"(\d{4})",s)[0])
        return res


    def get_kvalification(self):
        res = "None"
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'КВАЛИФИКАЦИЯ:',s):
                tmp = s.split(":")[1].replace(" ","").lower()
                if re.search(r'бакалавр',tmp):
                    res = 'bachelor'
                elif re.search(r'специалист', tmp):
                    res = 'specialist'
                elif re.search(r'магистр', tmp):
                    res = 'master'
                elif re.search(r'инженер', tmp):
                    res = 'engineer'
        return res

    def get_profiles(self):
        res = []
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'ПРОФИЛЬ:',s):
                tmp = s.split(":")[1]
                # if self.num_profiles==1:#one profile
                #     res.append(remove_quotes(tmp[1].lstrip()).capitalize())
                # else:#more profiles
                #     id = 0
                #     if tmp[1]:
                #         res.append(remove_quotes(tmp[1].lstrip()).capitalize())
                #     if self.type_edu == TypeEduPlan.FULLTIME:
                #         id=1
                #     elif self.type_edu==TypeEduPlan.EXTRAMURAL:
                #         id=1
                #     elif self.type_edu==TypeEduPlan.PART_TIME:
                #         id=0
                #     for j in range(1,self.num_profiles+id):
                #         res.append(remove_quotes(str(self.data_title[i+j]).lstrip()).capitalize())
                for p in tmp.split(","):
                    res.append(remove_quotes(p.strip()).capitalize())
        return res


    def get_srok_obucheniya(self):
        res = "0"
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'СРОК ОБУЧЕНИЯ:',s):
                res = re.findall(r"\d",s.split(":")[1])
        return res

    def get_form_education(self):#форма обучения
        res = "None"
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'ФОРМА ОБУЧЕНИЯ:', s):
                res = str(s.split(":")[1]).lstrip().capitalize()
            elif re.search(r'ФОРМА\s*ОБУЧЕНИЯ', s):
                res = str(s.split(" ")[0]).lstrip().capitalize()
        return res


    def get_program_education(self):#программа академического или прикладного бакалавра
        res = "None"
        for i in range(len(self.data_title)):
            s = str(self.data_title[i])
            if re.search(r'Программа', s):
                tmp = s.split("Программа")[1].lstrip()
                if re.search(r'приклад',tmp):
                    res = 'Applied'
                elif re.search(r'академ', tmp):
                    res = 'Academic'
        return res

    def get_names_disciplines(self):
        res = []

        for i in range(self.rows[0], self.rows[1]):
            if str(self.ws.cell(row=i, column=1).value).isdigit():
                tmp = str(self.ws.cell(row=i, column=self.col_name_discipline).value).split(' или ')
                for j in tmp:
                    if j and not j.isdigit():
                        res.append(j.lstrip().capitalize())  # Название дисциплины

        return remove_dublicates(res)

    def isProfilesCompare(self,current_row):
        prof = self.get_profiles()
        res = ""
        for j in range(1,5):
           if re.search(r'профиль',str(self.ws.cell(row=current_row, column=j).value).lower()):
               x = self.ws.cell(row=current_row, column=j).value
               for p in prof:
                if re.search(str(p).lower(), str(x).lower()): #профиль содержится в строке Х (работает для профилей записанных без скобок)
                    #print("Found: "+str(p)+"   "+str(x))
                    res = p
                else:
                   tmp = re.findall(r"\(\w*\)",p) #находим все слова в скобках т.к. используются сокращения
                   for k in tmp:
                    kk = k.replace('(','').replace(')','') #удаление скобок
                    r = re.search(str(kk).lower(), str(x).lower())
                    if r:
                        #print("Found: "+str(p) + "   " + str(r))
                        res = p
        return res

    def import_plan_of_dis_fulltime(self,ids, direction, weeks_in_semestrs, base_part_row): #добавление учебного плана для очной формы обучения

        current_profile = self.get_profiles()
        for i in range(base_part_row, self.rows[1]):
             if str(self.ws.cell(row=i, column=1).value).isdigit():
                 code_opop = self.ws.cell(row=i, column=ids[0]).value              #Код ОПОП
                 names = str(self.ws.cell(row=i, column=ids[1]).value).split(" или ")                   #Название дисциплины
                 competense = re.split('/|или',str(self.ws.cell(row=i, column=ids[16]).value))  # Коды формируемых компетенций

                 for nm in range(len(names)):
                     #ищим дисциплину names[j] из базы
                     total_hour = self.ws.cell(row=i, column=ids[2]).value             #Трудоемкость
                     zachotnih_edinic = self.ws.cell(row=i, column=ids[3]).value       #Зачетных единиц
                     audit_hour_sum   = int(self.ws.cell(row=i, column=ids[4]).value)  #Сумма аудиторных часов
                     audit_hour_lec   = int(self.ws.cell(row=i, column=ids[5]).value)  #Лекций, ч.
                     audit_hour_prakt = int(self.ws.cell(row=i, column=ids[6]).value)  #Практических работ, ч.
                     audit_hour_labs  = int(self.ws.cell(row=i, column=ids[7]).value)  #Лабораторных работ, ч.
                     samost_hour_sum  = int(self.ws.cell(row=i, column=ids[8]).value) #Всего самостоят раб., час
                     samost_hour_without_prepod = float(self.ws.cell(row=i, column=ids[9]).value)  #Самост.раб. без преподавателя
                     samost_hour_with_stud      = float(self.ws.cell(row=i, column=ids[10]).value)  #Самост.раб. преподвателя со студентом
                     samost_hour_with_group     = float(self.ws.cell(row=i, column=ids[11]).value)  #Самост.раб. преподвателя с группой

                     exam = isNone(self.ws.cell(row=i, column=ids[12]).value)           #Экзамен в семестре
                     zachot = isNone(self.ws.cell(row=i, column=ids[13]).value)         #Зачет в семестре
                     semestrs = getSemestrs_2d(str(exam),str(zachot))

                     KPKR = isNone(self.ws.cell(row=i, column=ids[14]).value)           #КП/КР
                     interaktiv_form_hour = self.ws.cell(row=i, column=ids[15]).value   #Итого в интерактивной форме

                     weeks = []
                     for sem in semestrs.split(","):
                         if int(sem)<=len(weeks_in_semestrs):
                            weeks.append(weeks_in_semestrs[int(sem)-1])

                     if(self.isUpdate==True):#режим обновления учебных планов
                         pl = Plans.objects.filter(code_OPOP=code_opop,discipline=Discipline.objects.get(name=names[nm].lstrip()),
                                                  direction_id=direction, training_form= 'fulltime' if self.type_edu==TypeEduPlan.FULLTIME \
                                                  else 'extramural' if self.type_edu==TypeEduPlan.EXTRAMURAL \
                                                  else 'parttime' if self.type_edu==TypeEduPlan.PART_TIME  \
                                                  else 'None', qualif = self.get_kvalification(), training_program=self.get_program_education(),
                                                  year = self.get_year())

                         prof_new = [] #создание списка обьектов профилей
                         for prof in current_profile:
                            tmp = Profiles.objects.filter(name=prof, direction_id=direction) # direction=Directions.objects.get(code=self.get_direction()[0])
                            for j in tmp:
                                if j not in prof_new:
                                    prof_new.append(j)

                         for plan in pl:
                            if(list(plan.profile.all())==prof_new): #сравнение профилей
                                plan.semestr = semestrs
                                plan.trudoemkost_all = total_hour
                                plan.trudoemkost_zachot_edinic = zachotnih_edinic
                                plan.hours_audit_work_sum = audit_hour_sum
                                plan.hours_lectures = audit_hour_lec
                                plan.hours_pract = audit_hour_prakt
                                plan.hours_labs = audit_hour_labs
                                plan.hours_samost_work_sum = samost_hour_sum
                                plan.hours_samost_wo_lec = samost_hour_without_prepod
                                plan.hours_samost_w_lec_w_stud = samost_hour_with_stud
                                plan.hours_samost_w_lec_w_group = samost_hour_with_group
                                plan.exam_semestr = exam
                                plan.zachot_semestr = zachot
                                plan.kursovya_work_project = KPKR
                                plan.zanatiya_in_interak_forms_hours = interaktiv_form_hour
                                plan.comps = competense[nm if len(competense) > 1 else 0].strip()
                                plan.weeks_count_in_semestr = ",".join(weeks)
                                plan.save()
                                print("Обновлен уч. план для", "    ", plan)
                            else:
                                print("Нет соответсвующих учебных планов для обновления")
                     else: #режим создания планов
                          p = Plans(code_OPOP=code_opop,discipline=Discipline.objects.get(name=names[nm].lstrip()), direction_id=direction,
                                     training_form= 'fulltime' if self.type_edu==TypeEduPlan.FULLTIME \
                                     else 'extramural' if self.type_edu==TypeEduPlan.EXTRAMURAL \
                                     else 'parttime' if self.type_edu==TypeEduPlan.PART_TIME else 'None',
                                     qualif = self.get_kvalification(), training_program=self.get_program_education(),
                                     year = self.get_year(), semestr=semestrs, trudoemkost_all = total_hour, trudoemkost_zachot_edinic = zachotnih_edinic,
                                     hours_audit_work_sum = audit_hour_sum,
                                     hours_lectures = audit_hour_lec,
                                     hours_pract = audit_hour_prakt,
                                     hours_labs = audit_hour_labs,
                                     hours_samost_work_sum= samost_hour_sum,
                                     hours_samost_wo_lec = samost_hour_without_prepod,
                                     hours_samost_w_lec_w_stud =  samost_hour_with_stud,
                                     hours_samost_w_lec_w_group = samost_hour_with_group,
                                     exam_semestr = exam,
                                     zachot_semestr = zachot,
                                     kursovya_work_project = KPKR,
                                     zanatiya_in_interak_forms_hours = interaktiv_form_hour,
                                     comps = competense[nm if len(competense)>1 else 0].strip(),
                                     weeks_count_in_semestr = ",".join(weeks),
                                     ministerstvo=Ministerstvo.objects.get(name=self.ministerstvo),
                                     univer=Univercity.objects.get(name=self.univercity)
                                     )
                          p.save()
                          print(names[nm])

                          for prof in current_profile:
                              tmp = Profiles.objects.filter(name=prof, direction_id=direction) # direction=Directions.objects.get(code=self.get_direction()[0])
                              for j in tmp:
                                 p.profile.add(j)
             else:
                 tmp_profile = self.isProfilesCompare(current_row=i)
                 if (tmp_profile and str(current_profile)!=tmp_profile):
                    print("Change profile " + str(current_profile) + " to " + tmp_profile)
                    current_profile = [tmp_profile]

    def import_plan_of_dis_exramural(self,code_opop_id, discipline_name_id, total_id, zachot_edinic_id, audit_sum_id, \
                                     hour_lec_id, hour_prakt_id, hour_labs_id, exam_id, zachot_id, KR_id, \
                                     kursovoi_proekt_id, kursovoi_rabota_id,direction, base_part_row):
        current_profile = self.get_profiles()

        for i in range(base_part_row, self.rows[1]):
             if str(self.ws.cell(row=i, column=1).value).isdigit():
                 code_opop = self.ws.cell(row=i, column=code_opop_id).value              #Код ОПОП
                 names = str(self.ws.cell(row=i, column=discipline_name_id).value).split(" или ")                   #Название дисциплины
                 for nm in names:
                     #ищим дисциплину names[j] из базы
                     total_hour = self.ws.cell(row=i, column=total_id).value                     #Трудоемкость
                     zachotnih_edinic = self.ws.cell(row=i, column=zachot_edinic_id).value       #Зачетных единиц
                     audit_hour = [self.ws.cell(row=i, column=audit_sum_id).value,  #Сумма аудиторных часов
                                   self.ws.cell(row=i, column=hour_lec_id).value,  #Лекций, ч.
                                   self.ws.cell(row=i, column=hour_prakt_id).value,  #Практических работ, ч.
                                   self.ws.cell(row=i, column=hour_labs_id).value]  #Лабораторных работ, ч.
                     samost_hour = int(total_hour) - int(audit_hour[0]) #Всего самостоят раб., час

                     exam = isNone(self.ws.cell(row=i, column=exam_id).value)           #Экзамен в семестре
                     zachot = isNone(self.ws.cell(row=i, column=zachot_id).value)       #Зачет в семестре
                     semestrs = getSemestrs_2d(str(exam), str(zachot))

                     KP = isNone(self.ws.cell(row=i, column=KR_id).value)               #КР
                     if self.ws.cell(row=i, column=kursovoi_proekt_id).value: #Курсовой проект
                        KPKR = "{0}{1}".format(self.ws.cell(row=i, column=kursovoi_proekt_id).value, "КП")
                     elif self.ws.cell(row=i, column=kursovoi_rabota_id).value: #Курсовая работа
                        KPKR = "{0}{1}".format(self.ws.cell(row=i, column=kursovoi_rabota_id).value, "КР")
                     else:
                        KPKR = "-"

                     if (self.isUpdate == True):  # режим обновления учебных планов
                         pl = Plans.objects.filter(code_OPOP=code_opop, direction_id=direction,
                                   discipline=Discipline.objects.get(name=nm.lstrip()),
                                   training_form='extramural' if self.type_edu == TypeEduPlan.EXTRAMURAL \
                                   else 'parttime' if self.type_edu == TypeEduPlan.PART_TIME else 'None',
                                   qualif=self.get_kvalification(), training_program=self.get_program_education(),
                                   year=self.get_year())

                         prof_new = []  # создание списка обьектов профилей
                         for prof in current_profile:
                             tmp = Profiles.objects.filter(name=prof,direction_id=direction)
                             for j in tmp:
                                 if j not in prof_new:
                                     prof_new.append(j)

                         for plan in pl:
                             if (list(plan.profile.all()) == prof_new):  # сравнение профилей
                                 plan.semestr = semestrs
                                 plan.trudoemkost_all = total_hour
                                 plan.trudoemkost_zachot_edinic = zachotnih_edinic
                                 plan.hours_audit_work_sum = audit_hour[0]
                                 plan.hours_lectures = audit_hour[1]
                                 plan.hours_pract = audit_hour[2]
                                 plan.hours_labs = audit_hour[3]
                                 plan.hours_samost_work_sum = samost_hour
                                 plan.hours_samost_wo_lec = samost_hour
                                 plan.hours_samost_w_lec_w_stud = 0
                                 plan.hours_samost_w_lec_w_group = 0
                                 plan.exam_semestr = exam
                                 plan.zachot_semestr = zachot
                                 plan.kursovya_work_project = KPKR
                                 plan.kontrolnaya_work = KP
                                 plan.save()
                                 print("Обновлен уч. план для", "    ", plan)
                             else:
                                 print("Нет соответсвующих учебных планов для обновления")
                     else:  # режим создания планов
                         p = Plans(code_OPOP=code_opop, direction_id=direction,
                                   discipline=Discipline.objects.get(name=nm.lstrip()),
                                   training_form='extramural' if self.type_edu == TypeEduPlan.EXTRAMURAL \
                                       else 'parttime' if self.type_edu == TypeEduPlan.PART_TIME else 'None',
                                   qualif=self.get_kvalification(), training_program=self.get_program_education(),
                                   year=self.get_year(), semestr=semestrs, trudoemkost_all=total_hour,
                                   trudoemkost_zachot_edinic=zachotnih_edinic,
                                   hours_audit_work_sum=audit_hour[0],
                                   hours_lectures=audit_hour[1],
                                   hours_pract=audit_hour[2],
                                   hours_labs=audit_hour[3],
                                   hours_samost_work_sum=samost_hour,
                                   hours_samost_wo_lec=samost_hour,
                                   hours_samost_w_lec_w_stud=0,
                                   hours_samost_w_lec_w_group=0,
                                   exam_semestr=exam,
                                   zachot_semestr=zachot,
                                   kursovya_work_project=KPKR,
                                   kontrolnaya_work=KP,
                                   ministerstvo=Ministerstvo.objects.get(name=self.ministerstvo),
                                   univer=Univercity.objects.get(name=self.univercity)
                                   )
                         p.save()
                         for prof in current_profile:
                             tmp = Profiles.objects.filter(name=prof,direction_id=direction) #Directions.objects.get(code=self.get_direction()[0])
                             for j in tmp:
                                p.profile.add(j)
                         print("{0} {1} {2}/{3}, {4}, {5}, exam={6} zachot={7} KPKR={8}, Контраб={9}".format(code_opop,nm,total_hour,isNone(zachotnih_edinic),audit_hour,samost_hour,exam,zachot,KPKR, KP))
             else:
                 tmp_profile = self.isProfilesCompare(current_row=i)
                 if (tmp_profile and str(current_profile)!=tmp_profile):
                    print("Change profile " + str(current_profile) + " to " + tmp_profile)
                    current_profile = [tmp_profile]

    def import_plan_of_discipline(self):

        directs = self.get_direction()

        if Directions.objects.filter(code=directs[0]).count()==0: #Если нет направления подготовки, то добавляем
            print('add direction ' + directs[1])
            tmp = Directions(name=str(directs[1]), code=str(directs[0]), number_prikaz=self.num_prikaz, date_prikaz=self.date_prikaz, deparmt=self.departament)
            tmp.save()
        if Profiles.objects.filter(direction__code=directs[0]).count()==0:
            for prof in self.get_profiles():
                print('add profile ' + prof)
                tmp = Profiles(direction=Directions.objects.get(code=directs[0]),name=prof)
                tmp.save()

        for d in self.get_names_disciplines():
            if Discipline.objects.filter(name=d).count()==0:
                print('add discipline ' + d)
                tmp = Discipline(name=d)
                tmp.save()

        num_cols = [1, 0]
        num_rows = [self.rows[0], 0]
        for j in range(1, self.ws.max_column):
            for i in range(self.rows[0], self.rows[1]):
                s = str(self.ws.cell(row=i, column=j).value)  #
                if re.search(r"кп\s*/\s*кр", s.lower()) or re.search(r"объем\s*работ,\s*час", s.lower()):
                    num_cols[1] = j+1
                    if self.type_edu != TypeEduPlan.FULLTIME:
                        num_cols[1] = j + 4
                    print("КП/КР  col=", j)
                    break
                elif re.search(r"базовая\s*часть", s.lower()):
                    num_rows[1] = i
                    print("базовая часть row=", i)
                    break
                elif re.search(r"интерактивной", s.lower()):
                    hours_interaktiv_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"коды формируемых компетенций", s.lower()):
                    code_kompetence_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"недель\s*в\s*семестре", s.lower()):
                    week_in_semestr_cell = [i, j]
                    print(str(j) + "  " + s)

        tmp_bez_prepod = ["",0]
        kursovoi_flag=False

        for j in range(num_cols[0],num_cols[1]):
            for i in range(num_rows[0],num_rows[1]):
                s = str(self.ws.cell(row=i, column=j).value)
                if re.search(r"код\s*опоп", s.lower()) or re.search(r"опоп", s.lower()):
                    code_opop_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"наименование",s.lower()):#Название дисциплины
                    discipline_name_id = j
                    print(str(j) + "  " + s)
                elif (re.search(r"час", s) and not re.search(r"Объем работ, час",s)) or re.search(r"общие часы", s.lower()):  # Общая трудоемкость
                    total_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"зет",s.lower()):#Зачетных единиц
                    zachot_edinic_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"Экз.",s) or re.search(r"экзамен",s):  #Экзамен в семестре
                    exam_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"Зач.",s) or re.search(r"зачет",s): #Зачет в семестре
                    zachot_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"КР",s) or (re.search(r"КП/КР",s) and self.type_edu==TypeEduPlan.FULLTIME):#Контрольная работа
                    KPKR_id = j
                    KR_id = j
                    print(str(j) + " " + s)
                elif re.search(r"Курсовой",s):#Курсовой проект или курсовая работа
                    kursovoi_flag = True
                elif re.search(r"П",s) and kursovoi_flag:#Курсовой проект
                    kursovoi_proekt_id = j
                    print(str(j) + " " + s)
                elif re.search(r"Р",s) and kursovoi_flag:#Курсовая работа
                    kursovoi_rabota_id = j
                    kursovoi_flag = False
                    print(str(j) + "  " + s)
                elif re.search(r"Объем работ,\s*час",s) or re.search(r"аудиторная\s*работа",s) or re.search(r"контактная\s*работа", s):#Всего аудиторных часов
                    audit_sum_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"лек.",s.lower()):#Лекций,час.
                    hour_lec_id = j
                    print(str(j) + "  " + s)
                elif re.search(r"Пр",s) or re.search(r"прак",s):#Практических работ,час.
                    hour_prakt_id = j
                    print(str(j) + " " + s)
                elif re.search(r"Лаб",s) or re.search(r"лаб.",s):#Лабораторных работ,час.
                    hour_labs_id = j
                    print(str(j) + "  " + s)
                elif self.type_edu==TypeEduPlan.FULLTIME:
                    if re.search(r"самостоят. работа", s):  # Всего объем самостоятельной работы, часов
                        samost_sum_id = j
                        print(str(j) + "  " + s)
                    elif  re.search(r"студ.",s):
                        samost_with_stud_id = j
                        print(str(j) + "  " + s)
                    elif re.search(r"гр.", s):
                        samost_with_group_id = j
                        print(str(j) + "  " + s)
                    elif re.search(r"без", s):
                        tmp_bez_prepod[0] = tmp_bez_prepod[0] + s
                        tmp_bez_prepod[1] = j
                    elif re.search(r"пре-", s):
                        tmp_bez_prepod[0] = tmp_bez_prepod[0] + " "+s[0:3]
                        tmp_bez_prepod[1] = j if (tmp_bez_prepod[1] % j)==0 else 0xFF
                    elif re.search(r"по-", s):
                        tmp_bez_prepod[0] = tmp_bez_prepod[0] + s[0:2]
                        tmp_bez_prepod[1] = j if (tmp_bez_prepod[1] % j) == 0 else 0xFF
                    elif re.search(r"дав.", s):
                        tmp_bez_prepod[0] = tmp_bez_prepod[0] + s[0:3]
                        tmp_bez_prepod[1] = j if (tmp_bez_prepod[1] % j) == 0 else 0xFF

        if self.type_edu==TypeEduPlan.FULLTIME:
            week_in_semestr = []
            for i in range(week_in_semestr_cell[1], self.ws.max_column):
                 s = str(self.ws.cell(row=week_in_semestr_cell[0]+1, column=i).value)
                 if s.isdigit():
                     if re.search(r"\d+", s.lower()):
                         print(str(i) + "  " + s)
                         week_in_semestr.append(s)
            if tmp_bez_prepod[1] != 0xFF:
                samost_without_prepod_id = tmp_bez_prepod[1]

            self.import_plan_of_dis_fulltime(ids = [code_opop_id,discipline_name_id,total_id,zachot_edinic_id,
                                                     audit_sum_id, hour_lec_id, hour_prakt_id, hour_labs_id,
                                                     samost_sum_id, samost_without_prepod_id, samost_with_stud_id,
                                                     samost_with_group_id, exam_id, zachot_id, KPKR_id,
                                                     hours_interaktiv_id, code_kompetence_id], direction=Directions.objects.get(code=directs[0]).id,
                                                     weeks_in_semestrs=week_in_semestr, base_part_row = num_rows[1])
        elif self.type_edu==TypeEduPlan.EXTRAMURAL or self.type_edu==TypeEduPlan.PART_TIME:
            self.import_plan_of_dis_exramural(code_opop_id = code_opop_id,
                                               discipline_name_id = discipline_name_id,
                                               total_id=total_id,
                                               zachot_edinic_id=zachot_edinic_id,
                                               audit_sum_id=audit_sum_id,
                                               hour_lec_id=hour_lec_id,
                                               hour_prakt_id=hour_prakt_id,
                                               hour_labs_id=hour_labs_id,
                                               exam_id=exam_id,
                                               zachot_id=zachot_id,
                                               KR_id=KR_id,
                                               kursovoi_proekt_id=kursovoi_proekt_id,
                                               kursovoi_rabota_id=kursovoi_rabota_id,
                                               direction=Directions.objects.get(code=directs[0]).id,
                                               base_part_row = num_rows[1])

